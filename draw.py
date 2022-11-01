from openpyxl.styles import Alignment, PatternFill, Font, Border
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side
from openpyxl import Workbook

FILE_NAME = "episode_ratings.xlsx"
BG_HEX = "262626"  # background color

# color for cells and text depending on what episode and season it is
LOW_RATING_HEX = "B04A4A"  # light red
MID_RATING_HEX = "F5862B"  # orange
HIGH_RATING_HEX = "1E9A47"  # green
MAX_RATING_HEX = "31869B"  # turquoise

RATING_CELL_HEX = "262626"  # black-ish
RATING_FONT_HEX = "FFFFFF"  # white
CELL_WIDTH = 5  # width between each cell on the x axis (columns)
BORDER = True
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"))


def fill_cell(cell, value):
    value = float(value)
    if value < 7:
        cell.fill = PatternFill("solid", start_color=LOW_RATING_HEX)
    elif value < 8:
        cell.fill = PatternFill("solid", start_color=MID_RATING_HEX)
    elif value < 10:
        cell.fill = PatternFill("solid", start_color=HIGH_RATING_HEX)
    elif value == 10:
        cell.fill = PatternFill("solid", start_color=MAX_RATING_HEX)


def draw_spreadsheet(ratings, amount):
    wb = Workbook()
    ws = wb.active

    # col and row starts at 2 because the first ones are to be left untouched
    col_count = 2
    for season in ratings:
        row_count = 2
        cell = ws.cell(row=1, column=season + 1, value=season)
        cell.alignment = Alignment(horizontal="center")  # centers cells
        cell.fill = PatternFill("solid", start_color=RATING_CELL_HEX)
        cell.font = Font(color=RATING_FONT_HEX)
        for episode in ratings[season]:
            cell = ws.cell(row=row_count, column=col_count, value=episode)
            cell.alignment = Alignment(horizontal="center")
            fill_cell(cell, episode)
            row_count += 1
        col_count += 1

    # find season with most episodes
    episodes_list = [len(ratings[season]) for season in ratings]
    max_episodes = max(episodes_list)

    # put the correct number of episodes on the columns
    for i in range(max_episodes + 1):
        cell = ws.cell(row=i + 1, column=1, value=i)
        cell.alignment = Alignment(horizontal="center")  # centers rest of cells
        cell.fill = PatternFill("solid", start_color=RATING_CELL_HEX)
        cell.font = Font(color=RATING_FONT_HEX)
    ws.cell(row=1, column=1, value="")  # clears first cell (A1)

    # sets the correct width for each cell
    dim_holder = DimensionHolder(worksheet=ws)
    for col in range(ws.min_column, ws.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(
            ws, min=col, max=col, width=CELL_WIDTH)
    ws.column_dimensions = dim_holder

    # formats the rest: background color and border
    for col in range(amount + 1):
        for row in range(max_episodes + 1):
            cell = ws.cell(row=row + 1, column=col + 1)
            fill = str(cell.fill)[136:144]  # extracts bg color from cell
            if fill == "00000000":  # fills bg color if the episode doesnt exist
                cell.fill = PatternFill("solid", start_color=BG_HEX)
            if BORDER == True:
                cell.border = THIN_BORDER

    wb.save(filename=FILE_NAME)
    print("\nEverything is done",
        f"— you can now close this window and open up '{FILE_NAME}'")